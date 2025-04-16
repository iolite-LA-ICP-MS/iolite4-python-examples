"""
A python-based export script for iolite 4 has no expectations or functions
that will be automatically run. It is simply evaulated when used as an export script.

Before evaluating the script a few additional objects are added to the module:

data    an interface to iolite's C++ data. E.g. you can get
        existing time series data or selection groups with it
        as well as make new ones. Importantly, you can get
        results via data.result(selection, channel)

IoLog   an interface to iolite's logging facility. You can add
        messages with, e.g., IoLog.debug('My message')

Using an export script you could, e.g., export data in a custom format and/or order, or
write results to a remote database.

As an example here, we'll write our results in the format of the plasmage.org data
reporting template: http://www.plasmage.org/recommendations/home.html
"""

from datetime import datetime
import numpy as np
import pandas as pd
from functools import partial
#from xlwt import Workbook, easyxf
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side

# Make sure file name ends with .xlsx
export_filepath = data.exportFilePath()
if not export_filepath.endswith('.xlsx'):
    export_filepath = export_filepath + '.xlsx'

# Insert your Organisation Details below
author_details = datetime.now().strftime("%d %B %Y") + ', School of Earth Sciences, The University of Melbourne'

wb = Workbook()
#ws = wb.add_sheet('Table', cell_overwrite_ok=True)
ws = wb.active
ws.title = 'U-Pb results'

sel_grp_list = data.selectionGroupList(data.ReferenceMaterial) + data.selectionGroupList(data.Sample)

sels = [s for sg in sel_grp_list for s in sg.selections()]

class ChannelNames:
    Pb206_cps = data.timeSeriesNames(data.Intermediate, {'Mass': 206, 'DRSType': 'BaselineSubtracted'})[0]
    Uppm = 'Approx_U_PPM'
    Th_U = 'Th/U'
    Pb206_Pb204 = 'Final Pb206/Pb204'
    Pb206_U238 = 'Final Pb206/U238'
    Pb206_U238_age = 'Final Pb206/U238 age'
    Pb207_U235 = 'Final Pb207/U235'
    Pb207_U235_age = 'Final Pb207/U235 age'
    Pb208_Pb206 = 'Final Pb208/Pb206'
    U238_Pb206 = 'Final U238/Pb206'
    Pb207_Pb206 = 'Final Pb207/Pb206'
    Pb207_Pb206_age = 'Final Pb207/Pb206 age'
    Wetherill_rho = 'rho 206Pb/238U v 207Pb/235U'
    TW_rho = 'rho 207Pb/206Pb v 238U/206Pb'
    Pb208_Th232 = 'Final Pb208/Th232'
    Pb208_Th232_age = 'Final Pb208/Th232 age'


def channel_data(channel_name, selection):
    try:
        result = data.result(selection, data.timeSeries(channel_name))
    except:
        return ('#N/A', '#N/A', '#N/A', '#N/A')

    try:
        u1s_pct = 0.5*100*result.uncertaintyAs2SE()/result.value()
        u2s_abs = result.uncertaintyAs2SE()
        pu2s_abs = result.propagatedUncertainty()
        return (result.value(), u1s_pct, u2s_abs, pu2s_abs)
    except ZeroDivisionError:
        return ('#N/A', '#N/A', '#N/A', '#N/A')


def selection_data(property_name, selection):
    value = selection.property(property_name)
    if not value:
        value = selection.group().name
    return (value,)


def associated_data(result_name, selection):
    return (data.associatedResult(selection, result_name).value(),)


def th_u_ratio(selection):
    try:
        return (1/data.result(selection, data.timeSeries("Final U/Th")).value(),)
    except:
        return('#N/A',)

def pb206_204(selection):
    try:
        Pb206 = data.timeSeriesList(data.Intermediate, {'Mass': 206, 'DRSType': 'BaselineSubtracted'})[0].dataForSelection(selection)
        Pb204 = data.timeSeriesList(data.Intermediate, {'Mass': 204, 'DRSType': 'BaselineSubtracted'})[0].dataForSelection(selection)
        ratio = Pb206/Pb204
        mean_ratio = np.nanmean(ratio)
        pct1se = 100*(np.nanstd(ratio)/np.sqrt(len(ratio)))/mean_ratio
        return (np.nanmean(ratio), pct1se)
    except Exception as e:
        print(e)
        return ('#N/A', '#N/A')
        
def pb208_206(selection):
    try:
        Pb208 = data.timeSeriesList(data.Intermediate, {'Mass': 208, 'DRSType': 'BaselineSubtracted'})[0].dataForSelection(selection)
        Pb206 = data.timeSeriesList(data.Intermediate, {'Mass': 206, 'DRSType': 'BaselineSubtracted'})[0].dataForSelection(selection)        
        ratio = Pb208/Pb206
        mean_ratio = np.nanmean(ratio)
        pct1se = 100*(np.nanstd(ratio)/np.sqrt(len(ratio)))/mean_ratio
        return (np.nanmean(ratio), pct1se)
    except Exception as e:
        print(e)
        return ('#N/A', '#N/A')

def f206c(selection):
    '''
    From Horstwood et al. (2016)
    f206c = (64model/64meas)*100
    where
    64model = 206/204 from Pb evolution model at uncorrected 7/6 age
    64meas = 206/204 measured in sample
    '''
    try:
        c64 = lambda a: 0.023*(a/1e3)**3 - 0.359*(a/1e3)**2 - 1.008*(a/1e3) + 19.04

        age76 = data.timeSeries('Final Pb207/Pb206 age').dataForSelection(selection)
        Pb206 = data.timeSeriesList(data.Intermediate, {'Mass': 206, 'DRSType': 'BaselineSubtracted'})[0].dataForSelection(selection)
        Pb204 = data.timeSeriesList(data.Intermediate, {'Mass': 204, 'DRSType': 'BaselineSubtracted'})[0].dataForSelection(selection)
        r64 = Pb206/Pb204
        
        f206c = 100*c64(age76)/r64
        mean_f206c = np.nanmean(f206c)
        return (mean_f206c,)
    except Exception as e:
        print(e)
        return ('#N/A',)

def conc_pct(selection):
    try:
        age6_38 = data.result(selection, data.timeSeries(ChannelNames.Pb206_U238_age)).value()
        age7_6 = data.result(selection, data.timeSeries(ChannelNames.Pb207_Pb206_age)).value()
        return (100*age6_38/age7_6,)
    except ZeroDivisionError:
        return ('#N/A', '#N/A', '#N/A', '#N/A')


def write_header():
    # write the PlasmaAge template header
    c = ws.cell(row=1, column=1, value=author_details)
    c.font = Font(bold=True, size=14)
    c = ws.cell(row=1, column=9, value='Data for Tera-Wasserburg plot')
    c.font = Font(bold=True, size=14)
    c = ws.cell(row=1, column=16, value='Data for Wetherill plot')
    c.font = Font(bold=True, size=14)
    c = ws.cell(row=1, column=25, value='Dates')
    c.font = Font(bold=True, size=14)

    header_line2=['Identifier','Comments','f206c','206Pb','Uppm','Th/U','206Pb/204Pb','1s%','238U/206Pb','1s%','207Pb/206Pb','1s%','208Pb/206Pb','1s%','207Pb/235U','1s%','206Pb/238U','1s%','Rho','208Pb/232Th','1s%','207Pb/206Pb','2s (abs)','2ssys (abs)','206Pb/238U','2s (abs)','2ssys (abs)','207Pb/235U','2s (abs)','2ssys (abs)','208Pb/232Th','2s (abs)','2ssys (abs)','% conc']

    for col in range(1, len(header_line2)+1):
        ws.cell(row=2, column=col, value=header_line2[col-1])


def write_footer():
    last_row = ws.max_row + 2
    settings = data.dataReductionScheme('U-Pb Geochronology').settings()
    for key in settings.keys():
        ws.cell(row = last_row, column=1, value=key)
        ws.cell(row = last_row, column=2, value=settings[key])
        last_row += 1


def write_column(column_index, data_func=None, uncert_types=[], default_content='#N/A'):
    # iterate through selections and add column
    row_index = 3  # start at row 2 to leave room for header
    col_index_start = column_index

    for s in sels:
        column_index = col_index_start

        if data_func:
            ws.cell(row=row_index, column=column_index, value=data_func(s)[0])

            for i in uncert_types:
                column_index += 1
                ws.cell(row=row_index, column=column_index, value=data_func(s)[i])
        else:
            ws.cell(row=row_index, column=column_index, value=default_content)

        row_index += 1


def add_borders():
    def set_borders(sides, min_row, max_row, min_col, max_col):
        thick_border = Side(border_style='thick', color='000000')
        top_s = Side()
        left_s = Side()
        right_s = Side()
        bottom_s = Side()

        if 'top' in sides:
            top_s = thick_border

        if 'left' in sides:
            left_s = thick_border

        if 'right' in sides:
            right_s = thick_border

        if 'bottom' in sides:
            bottom_s = thick_border

        for row in ws.iter_rows(min_row, max_row, min_col, max_col):
            for cell in row:
                cell.border = Border(top=top_s, left=left_s, right=right_s, bottom=bottom_s)

    set_borders(['top'], 1, 1, 1, 34)
    set_borders(['bottom'], 2, 2, 1, 34)
    set_borders(['bottom'], ws.max_row, ws.max_row, 1, 34)
    for col in [8, 12, 14, 19, 21, 30, 33, 34]:
        set_borders(['right'], 1, ws.max_row, col, col)
        set_borders(['top', 'right'], 1, 1, col, col)
        set_borders(['bottom', 'right'], 2, 2, col, col)
        set_borders(['bottom', 'right'], ws.max_row, ws.max_row, col, col)


def set_number_formats():
    def set_fmt(format, col):
        for row in ws.iter_rows(3, ws.max_row, col, col):
            for cell in row:
                cell.number_format = format

    for col in [4,5,22,23,24,25,26,27,28,29,30]:
        set_fmt('0', col)

    for col in [6,10,16,18,34]:
        set_fmt('0.0', col)

    for col in [9,12,19]:
        set_fmt('0.00', col)

    for col in [15]:
        set_fmt('0.0000', col)

    for col in [11,17]:
        set_fmt('0.00000', col)

write_header()
write_column(1, data_func = partial(selection_data, 'Name'))
write_column(2, data_func = partial(selection_data, 'Comment'))
write_column(3, data_func = f206c)
write_column(4, data_func = partial(channel_data, ChannelNames.Pb206_cps))
write_column(5, data_func = partial(channel_data, ChannelNames.Uppm))
write_column(6, data_func = th_u_ratio)
write_column(7, data_func = pb206_204, uncert_types=[1])
write_column(9, data_func = partial(channel_data, ChannelNames.U238_Pb206), uncert_types=[1])
write_column(11, data_func = partial(channel_data, ChannelNames.Pb207_Pb206), uncert_types=[1])
write_column(13, data_func = pb208_206, uncert_types=[1])
write_column(15, data_func = partial(channel_data, ChannelNames.Pb207_U235), uncert_types=[1])
write_column(17, data_func = partial(channel_data, ChannelNames.Pb206_U238), uncert_types=[1])
write_column(19, data_func = partial(associated_data, ChannelNames.Wetherill_rho))
write_column(20, data_func = partial(channel_data, ChannelNames.Pb208_Th232), uncert_types=[1])
write_column(22, data_func = partial(channel_data, ChannelNames.Pb207_Pb206_age), uncert_types=[2,3])
write_column(25, data_func = partial(channel_data, ChannelNames.Pb206_U238_age), uncert_types=[2,3])
write_column(28, data_func = partial(channel_data, ChannelNames.Pb207_U235_age), uncert_types=[2,3])
write_column(31, data_func = partial(channel_data, ChannelNames.Pb208_Th232_age), uncert_types=[2,3])
write_column(34, data_func = conc_pct)
add_borders()
set_number_formats()
write_footer()

wb.save(export_filepath)
