#/ Type: Exporter
#/ Name: AGN Traces Exporter
#/ Authors: Bence Paul and Joe Petrus
#/ Description: An export script for creating an AGN formatted spreadsheet for importing into the AGN database.
#/ References: None
#/ Version: 0.1
#/ Contact: iolitesupport@icpms.com

import os
import subprocess

from shutil import copy2

from openpyxl import load_workbook

from iolite.Qt import Qt
from iolite.QtCore import QFile, QIODevice, QEventLoop
from iolite.QtGui import QComboBox, QHeaderView, QLabel, QLineEdit, QStackedWidget, QTableWidget, QTableWidgetItem, QToolButton, QVBoxLayout, QWidget 
from iolite.QtGui import QCheckBox
from iolite.QtUiTools import QUiLoader

'''
Globals: These are the default values for the script.

'''

TEMPLATE_PATH = '/Users/bence/Dropbox/testing/_AGN/Jan_2024_Template/GCDatapoint.template.xlsx'
UI_FILE_PATH = '/Users/bence/Library/CloudStorage/Dropbox/testing/_AGN/AGN_Exporter/'
USERS = ['Alan Grieg', 'Ashlea Wainwright', 'Bence Paul', 'Brandon Mahan', 'Janet Hergt', 'Jon Woodhead', 'Roland Maas']
LAB = 'Isotope Geochemistry,  The University of Melbourne'

MULTICOLLECTOR_MACHINE_NAMES = ['Thermo Nepture', 'Neoma', 'Nu Plasma II/III']
TOF_MACHINE_NAMES = ['Vitesse Text', 'icpTOF']

class SettingsWidget(QWidget):

    def __init__(self, parent=None):
        super().__init__(parent)

        self.setAttribute(Qt.WA_DeleteOnClose, True)

        self.setLayout(QVBoxLayout())

        self.ui_file = QFile(UI_FILE_PATH + "AGN_Exporter_Traces.ui")

        if not self.ui_file.open(QIODevice.ReadOnly):
            print(os.path.abspath(__file__) + "/test.ui")
            raise RuntimeError('Could not load settings ui')

        ui = QUiLoader().load(self.ui_file, self)
        self.layout().addWidget(ui)
        self.layout().setContentsMargins(0, 0, 0, 0)

        self.stackedWidget = ui.findChild(QStackedWidget, 'stackedWidget')
        self.cancelButton = ui.findChild(QToolButton, 'cancelButton')
        self.continueButton = ui.findChild(QToolButton, 'continueButton')
        self.backButton = ui.findChild(QToolButton, 'backButton')
        self.stepLabel = ui.findChild(QLabel, 'stepLabel')

        self.userComboBox = ui.findChild(QComboBox, 'userComboBox')
        self.labLineEdit = ui.findChild(QLineEdit, 'labLineEdit')
        self.litLineEdit = ui.findChild(QLineEdit, 'litLineEdit')
        self.fundingLineEdit = ui.findChild(QLineEdit, 'fundingLineEdit')
        self.scaleComboBox = ui.findChild(QComboBox, 'scaleComboBox')
        self.techniqueComboBox = ui.findChild(QComboBox, 'techniqueComboBox')
        self.uncertTypeComboBox = ui.findChild(QComboBox, 'uncertaintyTypeComboBox')
        self.trimSigFigsCheckBox = ui.findChild(QCheckBox, 'trimSigFigsCheckBox')
        self.sessionIDComboBox = ui.findChild(QComboBox, 'sessionIDComboBox')
        self.groupsTable = ui.findChild(QTableWidget, 'groupsTableWidget')
        self.duplicatesLabel = ui.findChild(QLabel, 'duplicatesLabel')
        self.duplicatesTable = ui.findChild(QTableWidget, 'duplicatesTableWidget')

        self.scaleComboBox.addItems(['Spot', 'Transect'])
        self.techniqueComboBox.addItems(['LA-ICP-MS (Trace Element Analysis)','LA-MC-ICP-MS','LA-ICP-TOF-MS'])
        self.userComboBox.addItems(USERS)
        self.labLineEdit.setText(LAB)
        self.uncertTypeComboBox.addItems(['2 standard error', '1σ', '2σ', '95%'])
        self.sessionIDComboBox.addItems([data.sessionUUID(), 'File Path'])

        '''
        Get primary RMs (calibrants) as these should not be included in exported results.
        '''
        calibrants = []

        for ch in data.timeSeriesList(data.Input):
            extStd = ch.property('External standard')
            if extStd is not None:
                calibrants.append(extStd)

        # Set up the groups table
        self.groupsTable.setColumnCount(3)
        self.groupsTable.setHorizontalHeaderLabels(['Group Name', 'Type', 'Selection Count'])
        self.groupsTable.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.groupsTable.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.groupsTable.setEditTriggers(QTableWidget.NoEditTriggers)
        def createGroupTableEntry(group, type):
            row = self.groupsTable.rowCount
            self.groupsTable.insertRow(row)

            groupNameItem = QTableWidgetItem(group.name)
            if group.name in calibrants:
                groupNameItem.setCheckState(Qt.Unchecked)
            else:
                groupNameItem.setCheckState(Qt.Checked)
            self.groupsTable.setItem(row, 0, groupNameItem)

            groupTypeItem = QTableWidgetItem(type)
            self.groupsTable.setItem(row, 1, groupTypeItem)

            selCountItem = QTableWidgetItem(str(group.count))
            selCountItem.setTextAlignment(Qt.AlignHCenter)
            self.groupsTable.setItem(row, 2, selCountItem)


        for group in data.selectionGroupList(data.ReferenceMaterial):
            createGroupTableEntry(group, 'Reference Material')

        for group in data.selectionGroupList(data.Sample):
            createGroupTableEntry(group, 'Sample')

        # Set up the duplicates table
        self.duplicatesTable.setColumnCount(2)
        self.duplicatesTable.setHorizontalHeaderLabels(['Element', 'Channel'])
        self.duplicatesTable.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.duplicatesTable.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.duplicatesTable.verticalHeader().hide() 
        self.duplicatesTable.setEditTriggers(QTableWidget.NoEditTriggers)
        elements = [ch.property('Element') for ch in data.timeSeriesList(data.Output)]
        duplicates = set([el for el in elements if elements.count(el) > 1])
        self.duplicate_options = []

        if len(duplicates) > 0:
            for el in duplicates:
                for ch in data.timeSeriesList(data.Output):
                    if ch.property('Element') == el:
                        self.duplicate_options.append(ch.name)

            for option in self.duplicate_options:
                ch = data.timeSeries(option)
                row = self.duplicatesTable.rowCount
                self.duplicatesTable.insertRow(row)
                self.duplicatesTable.setItem(row, 0, QTableWidgetItem(str(ch.property('Element'))))
                chItem = QTableWidgetItem(ch.name)
                chItem.setCheckState(Qt.Unchecked)
                self.duplicatesTable.setItem(row, 1, chItem)
        else:
            self.duplicatesTable.hide()
            self.duplicatesLabel.setText('No duplicate elements found. Continue to the next step.')

        self.backButton.clicked.connect(self.previousTab)
        self.continueButton.clicked.connect(self.nextTab)
        self.cancelButton.clicked.connect(lambda: self.close())

        self.stackedWidget.setCurrentIndex(0)
        self.stepLabel.setText(f'Step {self.stackedWidget.currentIndex+1} of 4')

    def previousTab(self):
        current_index = self.stackedWidget.currentIndex
        self.stackedWidget.setCurrentIndex(current_index-1)
        self.stepLabel.setText(f'Step {self.stackedWidget.currentIndex+1} of 4')
        self.continueButton.setText("Continue")

    def nextTab(self):
        current_index = self.stackedWidget.currentIndex
        # If we're already on the last tab, export the data
        if current_index == self.stackedWidget.count - 1:
            # Check that the user has selected at least one duplicate channel
            if self.duplicatesTable.isVisible() and len([ch for ch in self.duplicate_options if self.duplicatesTable.item(self.duplicate_options.index(ch), 1).checkState() == Qt.Checked]) == 0:
                self.duplicatesLabel.setStyleSheet('color: red')
                print('Please select at least one duplicate channel.')
                return
            self.continueButton.setText("Exporting...")
            self.exportData()
            return
        self.stackedWidget.setCurrentIndex(current_index+1)
        self.stepLabel.setText(f'Step {self.stackedWidget.currentIndex+1} of 4')
        if self.stackedWidget.currentIndex == self.stackedWidget.count - 1:
            self.continueButton.setText("Export Data")
        else:
            self.continueButton.setText("Continue")


    def exportData(self):
        print('Exporting data...')

        try:
            fp = copy2(TEMPLATE_PATH, export_filepath)
        except FileNotFoundError:
            print('Could not find template file')
            return
        except PermissionError:
            print('Permission denied to write to export file')
            return
        except:
            print('Could not create copy of template file')
            return

        wb = load_workbook(fp)
        conc_ws = wb['GC Concentrations']
        datap_ws = wb['GC Datapoints']

        # Work out which channels to process, removing unwanted duplicates from the duplicates table
        channels_to_process = list(data.timeSeriesNames(data.Output))
        for row in range(self.duplicatesTable.rowCount):
            if self.duplicatesTable.item(row, 1).checkState() == Qt.Unchecked:
                print('Removing ' + self.duplicatesTable.item(row, 1).text())
                channels_to_process.remove(self.duplicatesTable.item(row, 1).text())
        ts_list = [data.timeSeries(ch) for ch in channels_to_process]

        # Create dict of col numbers so that we don't have to look up which col
        # we need to insert the value every time
        col_dict = {}
        for ts in ts_list:
            el = ts.property('Element')

            colNo = 5
            while colNo < 359:
                # NOTE: Space after element name in spreadsheet. If this changes, the line below will fail to match
                if conc_ws.cell(3, colNo).value == el + ' ':
                    col_dict[el] = colNo
                    break

                colNo += 1

        # Get list of groups to export and their type
        groups_to_export = []
        for row in range(self.groupsTable.rowCount):
            grpName = self.groupsTable.item(row, 0).text()
            if self.groupsTable.item(row, 0).checkState() == Qt.Checked:
                groups_to_export.append(grpName)
        
        groupCounter = 5
        rowNo = 5
        for g in groups_to_export:
            print(f'Exporting group: {g}')
            group = data.selectionGroup(g)
            datapoint_name = group.name
            datap_ws.cell(groupCounter, 1, value=datapoint_name)
            if group.type == data.Sample:
                datap_ws.cell(groupCounter, 2, value=group.name)
            elif group.type == data.ReferenceMaterial:
                datap_ws.cell(groupCounter, 3, value=group.name)
            
            datap_ws.cell(groupCounter, 5, value = self.techniqueComboBox.currentText)
            if self.sessionIDComboBox.currentText == 'File Path':
                datap_ws.cell(groupCounter, 6, value = data.sessionFilePath())
            else:
                datap_ws.cell(groupCounter, 6, value = self.sessionIDComboBox.currentText)
            datap_ws.cell(groupCounter, 9, value = self.scaleComboBox.currentText)
            datap_ws.cell(groupCounter, 12, value = self.uncertTypeComboBox.currentText)
            datap_ws.cell(groupCounter, 15, value = self.userComboBox.currentText)
            datap_ws.cell(groupCounter, 14, value = self.labLineEdit.text)
            datap_ws.cell(groupCounter, 13, value = self.litLineEdit.text)
            datap_ws.cell(groupCounter, 16, value = self.fundingLineEdit.text)


            for sel in group.selections():
                conc_ws.cell(row=rowNo, column=1, value=datapoint_name)
                conc_ws.cell(row=rowNo, column=3, value=sel.name)

                for ts in ts_list:
                    if self.trimSigFigsCheckBox.isChecked():
                        tidy_string = data.result(sel, ts).tidy()
                        res_value = tidy_string.split('±')[0].strip()
                        res_uncert = tidy_string.split('±')[1].strip()
                    else:
                        res_value = data.result(sel, ts).valueInPPM()
                        res_uncert = data.result(sel, ts).uncertaintyAs2SE()

                    el = ts.property('Element')
                    conc_ws.cell(row=rowNo, column=int(col_dict[el]), value=float(res_value))
                    conc_ws.cell(row=rowNo, column=int(col_dict[el])+1, value=float(res_uncert))
                    conc_ws.cell(row=rowNo, column=int(col_dict[el])+2, value=float(ts.property('Mass')))

                rowNo += 1
            groupCounter += 1

        wb.save(fp)
        # Now use the system open the file in Excel
        subprocess.run(['open', export_filepath])
        self.close()

widget = SettingsWidget()
widget.show()
loop = QEventLoop()
widget.destroyed.connect(lambda: loop.quit)
loop.exec() # wait ...
print('finished')
